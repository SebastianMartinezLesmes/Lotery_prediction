from openpyxl import load_workbook
from collections import defaultdict
from datetime import datetime
import os

ARCHIVO_EXCEL = "resultados_astro.xlsx"

def cargar_datos_excel():
    if not os.path.exists(ARCHIVO_EXCEL):
        print("❌ Archivo Excel no encontrado.")
        return []

    wb = load_workbook(ARCHIVO_EXCEL, read_only=True)
    ws = wb.active

    encabezados = [cell.value for cell in ws[1]]
    indices = {nombre.lower(): i for i, nombre in enumerate(encabezados)}

    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        try:
            fecha = fila[indices.get("fecha")]
            if isinstance(fecha, str):
                fecha = datetime.strptime(fecha, "%Y-%m-%d")
            elif isinstance(fecha, datetime):
                fecha = fecha
            else:
                continue

            filas.append({
                "fecha": fecha,
                "lottery": fila[indices.get("lottery")],
                "result": str(fila[indices.get("result")]).zfill(4),
                "series": str(fila[indices.get("series")]).zfill(3)
            })
        except:
            continue

    return filas

def calcular_pesos(datos):
    hoy = datetime.today()
    pesos_resultado = defaultdict(lambda: defaultdict(float))
    pesos_serie = defaultdict(lambda: defaultdict(float))

    for fila in datos:
        dias_antiguedad = (hoy - fila["fecha"]).days + 1
        peso = 1 / dias_antiguedad

        loteria = fila["lottery"]
        pesos_resultado[loteria][fila["result"]] += peso
        pesos_serie[loteria][fila["series"]] += peso

    return pesos_resultado, pesos_serie

def predecir(pesos_resultado, pesos_serie):
    predicciones = {}

    for loteria in pesos_resultado:
        numero = max(pesos_resultado[loteria], key=pesos_resultado[loteria].get)
        serie = max(pesos_serie[loteria], key=pesos_serie[loteria].get)

        predicciones[loteria] = {
            "numero_probable": numero,
            "simbolo_probable": serie
        }

    return predicciones

def mostrar_predicciones(predicciones):
    print("\n🔮 Predicción basada en frecuencia ponderada (más peso a los resultados recientes):\n")
    for loteria, pred in predicciones.items():
        print(f"🎰 {loteria}")
        print(f"   Número más probable: {pred['numero_probable']}")
        print(f"   Símbolo más probable: {pred['simbolo_probable']}\n")

if __name__ == "__main__":
    datos = cargar_datos_excel()
    if datos:
        pesos_resultado, pesos_serie = calcular_pesos(datos)
        predicciones = predecir(pesos_resultado, pesos_serie)
        mostrar_predicciones(predicciones)
