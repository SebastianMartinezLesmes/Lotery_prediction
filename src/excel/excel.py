import os
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta
from src.api.API import obtener_resultados_históricos_astro
from src.excel.read_excel import obtener_ultima_fecha_excel
from src.utils.config import CREATE_DOC

def guardar_resultados_en_excel(nombre_archivo=CREATE_DOC):
    # Obtener la última fecha registrada en el archivo
    ultima_fecha = obtener_ultima_fecha_excel()
    if ultima_fecha:
        # Comenzar desde el día siguiente a la última fecha
        fecha_inicio = datetime.combine(ultima_fecha + timedelta(days=1), datetime.min.time())
        print(f"📅 Consultando desde {fecha_inicio.date()} hasta hoy.")
    else:
        fecha_inicio = None  # Si no hay fecha, se usará la predeterminada (2023-02-01)
        print("📅 No hay fechas previas. Consultando desde el inicio por defecto.")

    resultados_nuevos = obtener_resultados_históricos_astro(fecha_inicio)
    filas_nuevas = []

    columnas = ['fecha', 'lottery', 'slug', 'result', 'series']
    claves_existentes = set()

    # Cargar archivo si existe
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        ws = wb.active

        # Leer claves existentes para evitar duplicados
        for row in ws.iter_rows(min_row=2, values_only=True):
            clave = tuple(row[:5])  # Las primeras 5 columnas
            claves_existentes.add(clave)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(columnas)

    # Agregar nuevas filas si no son duplicadas
    for dia in resultados_nuevos:
        fecha = dia['fecha']
        for resultado in dia['resultados']:
            fila = [
                fecha,
                resultado.get('lottery'),
                resultado.get('slug'),
                resultado.get('result'),
                resultado.get('series')
            ]
            clave = tuple(fila)
            if clave not in claves_existentes:
                ws.append(fila)
                claves_existentes.add(clave)

    wb.save(nombre_archivo)
    print(f"📁 Archivo '{nombre_archivo}' actualizado correctamente.")

if __name__ == "__main__":
    guardar_resultados_en_excel()
