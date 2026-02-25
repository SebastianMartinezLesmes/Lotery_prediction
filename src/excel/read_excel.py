from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
from src.core.config import settings

ARCHIVO = str(settings.get_excel_path())

def fecha_ayer():
    return (datetime.today() - timedelta(days=1)).date()

def obtener_ultima_fecha_excel():
    if not os.path.exists(ARCHIVO):
        print("El archivo Excel no existe.")
        return None

    try:
        wb = load_workbook(ARCHIVO)  # ← sin read_only=True
        ws = wb.active

        fecha_col = None
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value and str(col[0].value).strip().lower() == "fecha":
                fecha_col = col
                break

        if not fecha_col:
            print("!! No se encontró una columna llamada 'fecha'.")
            return None

        fechas = []
        for cell in fecha_col[1:]:  # Omitir encabezado
            try:
                if isinstance(cell.value, datetime):
                    fechas.append(cell.value.date())
                elif isinstance(cell.value, str):
                    fechas.append(datetime.strptime(cell.value, "%d/%m/%Y").date())
            except:
                continue

        if fechas:
            return max(fechas)
        else:
            return None
    except Exception as e:
        print(f"ERROR al leer el archivo: {e}")
        return None

def obtener_loterias_disponibles():
    if not os.path.exists(ARCHIVO):
        print("ERROR: Archivo Excel no encontrado.")
        return []

    wb = load_workbook(ARCHIVO, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    loteria_idx = headers.index("lottery") if "lottery" in headers else -1

    if loteria_idx == -1:
        print("!! No se encontró la columna 'lottery'")
        return []

    loterias = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[loteria_idx] is not None:
            loterias.add(row[loteria_idx].strip())

    return sorted(list(loterias))

if __name__ == "__main__":
    ayer = fecha_ayer()
    ultima_fecha = obtener_ultima_fecha_excel()

    print(f"Fecha de ayer: {ayer}")
    print(f"Última fecha registrada: {ultima_fecha}")

    if ultima_fecha is None or ultima_fecha < ayer:
        print(">> Se necesita actualizar los resultados.")
        print("   Ejecuta: python main.py --collect")
    else:
        print("OK Los resultados ya están actualizados.")

    loterias = obtener_loterias_disponibles()
    print("Loterías disponibles en el Excel:")
    for lot in loterias:
        print(f"   - {lot}")
