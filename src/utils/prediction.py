import os
import time
import json
import joblib
import warnings
import numpy as np
import pandas as pd

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from src.core.config import settings
from src.utils.training_simple import entrenar_modelos_por_loteria
from src.features.feature_engineering import generar_features


ARCHIVO_EXCEL = str(settings.get_excel_path())
CARPETA_MODELOS = str(settings.MODELS_DIR)

warnings.filterwarnings("ignore")

# Mapeo de códigos a signos zodiacales
ZODIACO = [
    "ARI", "TAU", "GEM", "CAN",
    "LEO", "VIR", "LIB", "ESC",
    "SAG", "CAP", "ACU", "PIS"
]

def obtener_zodiaco(codigo):
    """Convierte código numérico a signo zodiacal (abreviación de 3 letras)."""
    try:
        return ZODIACO[int(codigo)]
    except:
        return str(codigo)


def buscar_mejor_modelo(loteria, tipo):
    carpeta = Path(settings.MODELS_DIR)
    loteria = loteria.lower().replace(" ", "_")
    modelos = list(carpeta.glob(f"*_{loteria}_{tipo}.pkl"))

    if not modelos:
        return None, None, None

    mejor_modelo = None
    mejor_accuracy = -1
    mejor_payload = None

    for ruta in modelos:
        try:
            payload = joblib.load(ruta)
            if isinstance(payload, dict):
                accuracy = payload.get("accuracy", 0)
            else:
                accuracy = 0

            if accuracy > mejor_accuracy:
                mejor_accuracy = accuracy
                mejor_modelo = ruta
                mejor_payload = payload

        except Exception:
            continue

    return mejor_modelo, mejor_payload, mejor_accuracy


def guardar_resultado(prediccion, modelo_usado=None, confianza=None):
    """
    Guarda la predicción en formato JSON dentro de data/results.json
    
    Args:
        prediccion: El resultado predicho (dict o str)
        modelo_usado: El modelo que generó la predicción
        confianza: Nivel de confianza (si aplica)
    """
    data_dir = settings.DATA_DIR
    output_file = data_dir / "results.json"
    
    entrada = {
        "timestamp": datetime.now().isoformat(),
        "resultado": prediccion,
        "modelo": modelo_usado,
        "confianza": confianza
    }
    
    # Leer resultados anteriores si existen
    datos = []
    if output_file.exists():
        try:
            with open(output_file, "r", encoding="utf-8") as f:
                datos = json.load(f)
        except json.JSONDecodeError:
            pass
    
    datos.append(entrada)
    
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(datos, f, indent=4, ensure_ascii=False)


def cargar_datos_excel():
    """Carga datos desde el archivo Excel."""
    if not os.path.exists(ARCHIVO_EXCEL):
        print("ERROR: Archivo Excel no encontrado.")
        return pd.DataFrame()

    wb = load_workbook(ARCHIVO_EXCEL, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1] if cell.value is not None]

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fila = dict(zip(headers, row))
        if all(k in fila and fila[k] is not None for k in ['fecha', 'lottery', 'result', 'series']):
            try:
                fecha = fila['fecha']
                if isinstance(fecha, str):
                    fecha = datetime.strptime(fecha, "%Y-%m-%d")
                result = int(fila['result'])
                series = str(fila['series'])

                data.append({
                    "fecha": fecha,
                    "lottery": fila['lottery'],
                    "result": result,
                    "series": series
                })
            except Exception:
                pass
    return pd.DataFrame(data)


def preparar_datos(df, loteria):
    """Prepara datos para una lotería específica."""
    df = df[df["lottery"].str.upper() == loteria.upper()].copy()
    df = df.sort_values("fecha")
    # Garantizar que fecha sea datetime (viene como date desde Neon)
    df["fecha"] = pd.to_datetime(df["fecha"])
    df["dia"]       = df["fecha"].dt.day
    df["mes"]       = df["fecha"].dt.month
    df["anio"]      = df["fecha"].dt.year
    df["dia_semana"]= df["fecha"].dt.weekday
    return df


def predecir_para_loteria(df, loteria):

    inicio = time.time()

    df_loteria = preparar_datos(df, loteria)

    if len(df_loteria) < settings.TRAINING_CONFIGURE["min_records"]:
        print(f"⚠️ Datos insuficientes para {loteria}: {len(df_loteria)} registros")
        return

    # =========================
    # BUSCAR MEJOR MODELO
    # =========================

    result_path, result_payload, result_score = buscar_mejor_modelo(loteria, "result")
    series_path, series_payload, series_score = buscar_mejor_modelo(loteria, "series")

    if not (result_path and series_path):
        print(
            f"\n❌ No se encontraron modelos entrenados para '{loteria}'.\n"
            f"   Ejecuta primero: python main.py --entrenar --lottery \"{loteria}\"\n"
            f"   No se puede predecir sin modelos previos."
        )
        return

    modelo_result = result_payload["model"]
    modelo_series = series_payload["model"]

    print(f"✓ Mejor modelo RESULT: {result_path.name} | Accuracy={result_score:.4f}")
    print(f"✓ Mejor modelo SERIES: {series_path.name} | Accuracy={series_score:.4f}")

    # =========================
    # GENERAR FEATURES
    # =========================

    X_train_df = generar_features(df_loteria)

    df_loteria = df_loteria.tail(len(X_train_df))

    X_train = X_train_df.values
    y_result = df_loteria["result"].values
    y_series = (
        df_loteria["series"]
        .astype(str)
        .str.upper()
        .astype("category")
        .cat.codes
        .values
    )

    # =========================
    # FEATURES PARA PREDICCION
    # =========================

    features = X_train_df.tail(1).values

    if hasattr(modelo_result, "n_features_in_"):
        if features.shape[1] != modelo_result.n_features_in_:
            raise ValueError(
                f"El modelo espera {modelo_result.n_features_in_} features "
                f"pero recibió {features.shape[1]}"
            )

    # =========================
    # PREDICCION NUMERO + SIGNO (Paso 6: top 3 con confianza)
    # =========================

    # Verificar compatibilidad de features
    if hasattr(modelo_result, "n_features_in_"):
        if features.shape[1] != modelo_result.n_features_in_:
            raise ValueError(
                f"El modelo espera {modelo_result.n_features_in_} features "
                f"pero recibió {features.shape[1]}"
            )

    # Modelo compuesto (4 dígitos) o modelo clásico
    from src.utils.training_simple import _ModeloCompuesto

    if isinstance(modelo_result, _ModeloCompuesto):
        top3_numeros = modelo_result.top3_numeros(features)
        pred_result  = top3_numeros[0][0]
        confianza    = top3_numeros[0][1]
    elif hasattr(modelo_result, "predict_proba"):
        pred_probs_result = modelo_result.predict_proba(features)[0]
        top3_idx     = np.argsort(pred_probs_result)[-3:][::-1]
        top3_numeros = [(int(modelo_result.classes_[i]), float(pred_probs_result[i])) for i in top3_idx]
        pred_result  = top3_numeros[0][0]
        confianza    = top3_numeros[0][1]
    else:
        pred_result  = int(modelo_result.predict(features)[0])
        top3_numeros = [(pred_result, None)]
        confianza    = None

    # Series / signo
    if hasattr(modelo_series, "predict_proba"):
        pred_probs_series = modelo_series.predict_proba(features)[0]
        top3_series_idx   = np.argsort(pred_probs_series)[-3:][::-1]
        top3_signos  = [(obtener_zodiaco(modelo_series.classes_[i]), float(pred_probs_series[i])) for i in top3_series_idx]
        pred_series  = modelo_series.classes_[top3_series_idx[0]]
    else:
        pred_series  = modelo_series.predict(features)[0]
        top3_signos  = [(obtener_zodiaco(pred_series), None)]

    signo = obtener_zodiaco(pred_series)

    resultado_final = {
        "loteria":      loteria,
        "numero":       int(pred_result),
        "serie":        signo,
        "top3_numeros": top3_numeros,
        "top3_signos":  top3_signos,
    }

    # =========================
    # MOSTRAR RESULTADO
    # =========================

    print("\n🎯 PREDICCIÓN")
    print(f"  Lotería : {loteria}")
    print(f"  Número  : {str(pred_result).zfill(4)}   Signo: {signo}")

    if confianza is not None:
        print(f"\n  Top 3 números:")
        for num, prob in top3_numeros:
            bar = "█" * max(1, int((prob or 0) * 20))
            print(f"    {str(num).zfill(4)}  {(prob or 0):6.2%}  {bar}")

        print(f"\n  Top 3 signos:")
        for sg, prob in top3_signos:
            bar = "█" * max(1, int((prob or 0) * 20))
            print(f"    {sg}  {(prob or 0):6.2%}  {bar}")

    guardar_resultado(
        resultado_final,
        modelo_usado="RandomForest",
        confianza=float(confianza) if confianza else None
    )

    duracion = time.time() - inicio
    print(f"⏱ Tiempo de predicción: {duracion:.2f}s")


def main(filtro_loteria=None):
    # ── Sincronizar con Neon antes de predecir ────────────────────
    try:
        from src.database.sync import synchronize_database
        synchronize_database(filtro_loteria=filtro_loteria)
    except Exception as e_sync:
        print(f"⚠️  Sincronización con Neon falló (usando datos locales): {e_sync}")

    # ── Cargar datos desde Neon ───────────────────────────────────
    df = None
    loterias = []

    try:
        from src.database.connection import NeonConnection
        from src.database.repository import LotteriaRepository

        conn = NeonConnection()
        repository = LotteriaRepository(conn)

        loterias_neon = ["ASTRO SOL", "ASTRO LUNA"]
        if filtro_loteria:
            loterias_neon = [l for l in loterias_neon if filtro_loteria.upper() in l.upper()]

        frames = []
        for lot in loterias_neon:
            try:
                df_lot = repository.get_all_results(lot)
                if not df_lot.empty:
                    frames.append(df_lot)
                    loterias.append(lot)
            except Exception as e_lot:
                print(f"⚠️  No se pudo cargar {lot} desde Neon: {e_lot}")

        if frames:
            df = pd.concat(frames, ignore_index=True)
            print(f"Datos cargados desde Neon PostgreSQL: {len(df)} registros")
        conn.close()
    except Exception as e_neon:
        print(f"⚠️  Neon no disponible, usando fallback Excel: {e_neon}")

    # ── Fallback a Excel ──────────────────────────────────────────
    if df is None or df.empty:
        df = cargar_datos_excel()
        if df.empty:
            print("!! No se pudieron cargar datos.")
            return
        # Obtener loterías desde el Excel
        loterias = sorted(df["lottery"].dropna().unique().tolist())

    if not loterias:
        print("!! No se encontraron loterías disponibles.")
        return

    print(f"\nLoterías detectadas: {loterias}")
    if filtro_loteria:
        loterias = [l for l in loterias if filtro_loteria.lower() in l.lower()]

    for loteria in loterias:
        print(f"\n>> Procesando: {loteria}")
        df_loteria = preparar_datos(df, loteria)
        if not df_loteria.empty:
            predecir_para_loteria(df_loteria, loteria)
        else:
            print(f"!! No hay suficientes datos para {loteria}")

if __name__ == "__main__":
    main()