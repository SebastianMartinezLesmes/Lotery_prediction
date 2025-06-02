import os
from openpyxl import Workbook, load_workbook
from API import obtener_resultados_históricos_astro

def guardar_resultados_en_excel(nombre_archivo="resultados_astro.xlsx"):
    resultados_nuevos = obtener_resultados_históricos_astro()
    filas_nuevas = []

    columnas = ['fecha', 'lottery', 'slug', 'result', 'series']
    claves_existentes = set()

    # Cargar archivo si existe
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        ws = wb.active

        # Leer claves existentes para evitar duplicados
        for row in ws.iter_rows(min_row=2, values_only=True):
            clave = tuple(row[:5])  # Las primeras 5 columnas
            claves_existentes.add(clave)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(columnas)

    # Agregar nuevas filas si no son duplicadas
    for dia in resultados_nuevos:
        fecha = dia['fecha']
        for resultado in dia['resultados']:
            fila = [fecha, resultado.get('lottery'), resultado.get('slug'), resultado.get('result'), resultado.get('series')]
            clave = tuple(fila)
            if clave not in claves_existentes:
                ws.append(fila)
                claves_existentes.add(clave)

    wb.save(nombre_archivo)
    print(f"📁 Archivo '{nombre_archivo}' actualizado correctamente.")

if __name__ == "__main__":
    guardar_resultados_en_excel()
